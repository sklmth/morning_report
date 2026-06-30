"""
经营分析结果导出为 Excel
生成多 Sheet 的分析报告，可供二次处理
"""

import io
from datetime import datetime
from typing import Optional, Union

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from analytics.db import get_connection
from analytics.analyzer.metrics import (
    get_score_structure, get_person_efficiency,
    get_risk_alerts, get_branch_compare, get_trend, get_overview
)
from analytics.analyzer.forecast import get_progress_forecast


# ── 样式辅助 ──────────────────────────────────────────────────────────────────

def _header_fill(color="1F4E79"):
    return PatternFill("solid", fgColor=color)


def _side():
    return Side(style="thin", color="CCCCCC")


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")
RED_FONT = Font(name="微软雅黑", size=10, color="C00000")
GREEN_FONT = Font(name="微软雅黑", size=10, color="375623")
ORANGE_FILL = PatternFill("solid", fgColor="FFE699")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")


def _write_header_row(ws, row: int, cols: list[str], col_widths: list[float] = None):
    """写入表头行"""
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]


def _write_data_row(ws, row: int, values: list, fills: list = None):
    """写入数据行"""
    for i, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        if fills and i <= len(fills) and fills[i - 1]:
            cell.fill = fills[i - 1]


# ── Sheet 生成函数 ─────────────────────────────────────────────────────────────

def _sheet_overview(wb: openpyxl.Workbook, month: str, data: dict):
    ws = wb.create_sheet("总览")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:F1")
    ws["A1"] = f"端州分公司政企部经营分析总览 — {month}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("月份", data.get("month", "-")),
        ("数据日期", data.get("latest_date", "-")),
        ("净增积分", data.get("net_pts", 0)),
        ("增量积分", data.get("inc_pts", 0)),
        ("总高套数", data.get("total_gaotao", 0)),
        ("人均激励(元)", data.get("avg_incentive", 0)),
        ("总激励(元)", data.get("total_incentive", 0)),
        ("快照数量", data.get("snapshot_count", 0)),
    ]
    _write_header_row(ws, 2, ["指标", "数值"], [20, 20])
    for i, (k, v) in enumerate(kpis, 3):
        _write_data_row(ws, i, [k, v])


def _sheet_score_structure(wb: openpyxl.Workbook, month: str, data: dict):
    ws = wb.create_sheet("积分结构分析")
    dz = data.get("duanzhou", {})
    headers = ["项目", "积分值", "说明"]
    _write_header_row(ws, 1, ["端州分公司积分结构分析"], ["60"])
    ws.merge_cells("A1:C1")
    ws["A1"].font = TITLE_FONT

    rows = [
        ("净增积分(全业务)", dz.get("net_pts", 0), "= 增量 - 到期 - 拆机 - 降值"),
        ("增量积分", dz.get("inc_pts", 0), ""),
        ("  基本面", dz.get("base_pts", 0), ""),
        ("    其中：移动", dz.get("base_mobile", 0), ""),
        ("    其中：宽带", dz.get("base_bb", 0), ""),
        ("    其中：固话", dz.get("base_phone", 0), ""),
        ("    其中：ITV", dz.get("base_itv", 0), ""),
        ("    其中：智家", dz.get("base_smart", 0), ""),
        ("  双线", dz.get("twin_pts", 0), ""),
        ("    其中：互专", dz.get("twin_inet", 0), ""),
        ("    其中：组网", dz.get("twin_net", 0), ""),
        ("  其他业务", dz.get("other_pts", 0), ""),
        ("--- 流失 ---", "", ""),
        ("基本面到期积分", dz.get("base_expire", 0), "负值表示到期流失"),
        ("基本面降值积分", dz.get("base_decline", 0), "负值表示价值降低"),
        ("基本面拆机积分", dz.get("base_churn", 0), "负值表示拆机损失"),
        ("双线降值积分", dz.get("twin_decline", 0), ""),
        ("双线拆机积分", dz.get("twin_churn", 0), ""),
    ]
    _write_header_row(ws, 2, headers, [25, 15, 30])
    for i, row_data in enumerate(rows, 3):
        fill = RED_FILL if row_data[0].startswith("---") else None
        _write_data_row(ws, i, list(row_data), [fill, fill, fill])

    health = data.get("health", {})
    r = len(rows) + 4
    ws.cell(r, 1, "健康度指标").font = TITLE_FONT
    ws.merge_cells(f"A{r}:C{r}")
    _write_header_row(ws, r + 1, ["指标", "当前值(%)", "预警阈值(%)"], [25, 15, 15])
    health_rows = [
        ("拆机占比", health.get("churn_ratio", 0), 30),
        ("降值占比", health.get("decline_ratio", 0), 25),
        ("到期积分占比", health.get("expire_ratio", 0), 40),
    ]
    for i, (k, v, thresh) in enumerate(health_rows, r + 2):
        fill = RED_FILL if float(v or 0) > thresh else GREEN_FILL
        _write_data_row(ws, i, [k, v, thresh], [None, fill, None])

    # 全市各县分对比
    all_d = data.get("all_districts", [])
    if all_d:
        r2 = r + 6 + len(health_rows)
        ws.cell(r2, 1, "全市各县分积分对比").font = TITLE_FONT
        ws.merge_cells(f"A{r2}:F{r2}")
        _write_header_row(ws, r2 + 1, ["县分", "净增积分", "增量积分", "基本面", "双线", "其他"],
                          [15, 12, 12, 12, 12, 12])
        for i, d in enumerate(all_d, r2 + 2):
            _write_data_row(ws, i, [
                d.get("district"), round(d.get("net_pts") or 0, 1),
                round(d.get("inc_pts") or 0, 1), round(d.get("base_pts") or 0, 1),
                round(d.get("twin_pts") or 0, 1), round(d.get("other_pts") or 0, 1),
            ])


def _sheet_progress(wb: openpyxl.Workbook, month: str, data: dict):
    ws = wb.create_sheet("完成进度预测")
    ws.cell(1, 1, f"完成进度预测 — {month}").font = TITLE_FONT
    ws.merge_cells("A1:J1")

    fc = data.get("pts_forecast", {})
    meta_rows = [
        ("数据截至", data.get("latest_date", "-")),
        ("时间进度", f"{data.get('time_progress', 0):.1f}%"),
        ("剩余天数", data.get("remaining_days", 0)),
        ("当前净增积分", fc.get("current", 0)),
        ("日均积分(实际)", fc.get("daily_avg_actual", 0)),
        ("预测月末净增积分", fc.get("projected_month_end", 0)),
    ]
    for i, (k, v) in enumerate(meta_rows, 2):
        ws.cell(i, 1, k).font = Font(bold=True, name="微软雅黑")
        ws.cell(i, 2, v)

    r = len(meta_rows) + 3
    headers = ["姓名", "高套数", "增量积分", "预测月末高套",
               "预测月末积分", "预计激励(元)", "129档", "169档", "199+档", "状态"]
    _write_header_row(ws, r, headers, [12, 10, 12, 12, 12, 14, 10, 10, 10, 8])
    for i, p in enumerate(data.get("person_progress", []), r + 1):
        status_fill = {"green": GREEN_FILL, "yellow": ORANGE_FILL, "red": RED_FILL}.get(p.get("status"), None)
        _write_data_row(ws, i, [
            p.get("name"), p.get("total_gaotao"), round(p.get("inc_pts") or 0, 1),
            p.get("projected_gaotao_month"), p.get("projected_pts_month"),
            p.get("predicted_incentive"), p.get("tier_129"), p.get("tier_169"), p.get("tier_199"),
            {"green": "达标", "yellow": "注意", "red": "预警"}.get(p.get("status"), "-"),
        ], [None] * 9 + [status_fill])


def _sheet_person_efficiency(wb: openpyxl.Workbook, month: str, data: dict):
    ws = wb.create_sheet("人员效能")
    ws.cell(1, 1, f"人员效能分析 — {month}").font = TITLE_FONT
    ws.merge_cells("A1:L1")

    # 营服人员效能
    headers = ["姓名", "中心", "角色", "预计激励(元)", "综合高套",
               "新装高套", "存量高套", "揽装积分", "激励积分",
               "FTTR", "合约", "移动"]
    _write_header_row(ws, 2, headers, [12, 16, 14, 14, 10, 10, 10, 12, 12, 8, 8, 8])
    for i, s in enumerate(data.get("staff_efficiency", []), 3):
        _write_data_row(ws, i, [
            s.get("name"), s.get("center"), s.get("role"),
            round(s.get("predicted_incentive") or 0, 0),
            round(s.get("total_gaotao") or 0, 1), round(s.get("new_gaotao") or 0, 1),
            round(s.get("stock_gaotao") or 0, 1), round(s.get("device_pts") or 0, 1),
            round(s.get("incentive_pts") or 0, 1),
            s.get("fttr"), s.get("contract"), s.get("mobile"),
        ])

    r2 = len(data.get("staff_efficiency", [])) + 4
    ws.cell(r2, 1, "CP对效能").font = TITLE_FONT
    ws.merge_cells(f"A{r2}:H{r2}")
    cp_headers = ["中心", "CP组", "营销人员", "装维人员", "CP目标", "实际积分", "完成率(%)", "积分缺口"]
    _write_header_row(ws, r2 + 1, cp_headers, [16, 8, 12, 12, 12, 12, 12, 12])
    for i, cp in enumerate(data.get("cp_pairs", []), r2 + 2):
        rate = cp.get("completion_rate") or 0
        fill = GREEN_FILL if rate >= 100 else (ORANGE_FILL if rate >= 80 else RED_FILL)
        _write_data_row(ws, i, [
            cp.get("center"), cp.get("cp_group"), cp.get("sales_name"), cp.get("install_name"),
            round(cp.get("cp_target") or 0, 1), round(cp.get("cp_pts_total") or 0, 1),
            rate, round(cp.get("cp_gap") or 0, 1),
        ], [None, None, None, None, None, None, fill, None])


def _sheet_risk(wb: openpyxl.Workbook, month: str, data: dict):
    ws = wb.create_sheet("存量风险预警")
    ws.cell(1, 1, f"存量风险预警 — {month}").font = TITLE_FONT
    ws.merge_cells("A1:E1")

    alerts = data.get("alerts", [])
    _write_header_row(ws, 2, ["预警级别", "类型", "当前值", "说明"], [12, 20, 12, 40])
    for i, a in enumerate(alerts, 3):
        fill = RED_FILL if a.get("level") == "red" else ORANGE_FILL
        _write_data_row(ws, i, [
            {"red": "🔴 红色", "orange": "🟡 黄色"}.get(a.get("level"), a.get("level")),
            a.get("type"), a.get("value"), a.get("desc"),
        ], [fill] * 4)

    if not alerts:
        ws.cell(3, 1, "暂无风险预警")

    # 历史趋势
    hist = data.get("historical_trend", [])
    if hist:
        r = max(len(alerts) + 4, 6)
        ws.cell(r, 1, "近期历史趋势").font = TITLE_FONT
        ws.merge_cells(f"A{r}:F{r}")
        _write_header_row(ws, r + 1, ["月份", "净增积分", "增量积分", "拆机积分", "降值积分", "到期积分"],
                          [12, 12, 12, 12, 12, 12])
        for i, h in enumerate(hist, r + 2):
            _write_data_row(ws, i, [
                h.get("month"), round(h.get("net_pts") or 0, 1),
                round(h.get("inc_pts") or 0, 1), round(h.get("base_churn") or 0, 1),
                round(h.get("base_decline") or 0, 1), round(h.get("base_expire") or 0, 1),
            ])


def _sheet_trend(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("历史趋势")
    ws.cell(1, 1, "端州分公司月度经营趋势").font = TITLE_FONT
    ws.merge_cells("A1:F1")

    pts = data.get("pts_trend", [])
    if pts:
        _write_header_row(ws, 2, ["月份", "净增积分", "增量积分", "基本面", "双线", "其他业务"],
                          [12, 12, 12, 12, 12, 12])
        for i, p in enumerate(pts, 3):
            _write_data_row(ws, i, [
                p.get("month"), round(p.get("net_pts") or 0, 1),
                round(p.get("inc_pts") or 0, 1), round(p.get("base_pts") or 0, 1),
                round(p.get("twin_pts") or 0, 1), round(p.get("other_pts") or 0, 1),
            ])

    gaotao = data.get("gaotao_trend", [])
    r2 = len(pts) + 4
    if gaotao:
        ws.cell(r2, 1, "高套发展趋势").font = TITLE_FONT
        ws.merge_cells(f"A{r2}:D{r2}")
        _write_header_row(ws, r2 + 1, ["月份", "新增高套", "存量高套", "参与人数"], [12, 12, 12, 12])
        for i, g in enumerate(gaotao, r2 + 2):
            _write_data_row(ws, i, [
                g.get("month"), round(g.get("new_gaotao") or 0, 1),
                round(g.get("stock_gaotao") or 0, 1), g.get("person_count"),
            ])


# ── 主入口 ────────────────────────────────────────────────────────────────────

def build_analysis_excel(month: str, output: Union[str, io.BytesIO]) -> None:
    """
    生成完整分析报告 Excel
    output: 文件路径字符串 或 BytesIO 对象
    """
    conn = get_connection()
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认空 sheet

        overview_data = get_overview(month, conn)
        score_data = get_score_structure(month, conn)
        progress_data = get_progress_forecast(month, conn)
        person_data = get_person_efficiency(month, conn)
        risk_data = get_risk_alerts(month, conn)
        trend_data = get_trend(12, conn)

        _sheet_overview(wb, month, overview_data)
        _sheet_score_structure(wb, month, score_data)
        _sheet_progress(wb, month, progress_data)
        _sheet_person_efficiency(wb, month, person_data)
        _sheet_risk(wb, month, risk_data)
        _sheet_trend(wb, trend_data)

        if isinstance(output, (str, bytes, os.PathLike)):
            wb.save(output)
        else:
            wb.save(output)
    finally:
        conn.close()
