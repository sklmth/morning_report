"""把统计 DataFrame 写成带样式的 Excel，风格对齐「早会五张表」。

样式要点（取自 assets/早会五张表.xlsx 的模板1）：
    字体      微软雅黑 11
    表头      主题色 accent1(4874CB) + tint 0.6 的浅蓝底、居中
    边框      全表细线
    完成率    百分比显示
    预约数    < 5 户标红（红字 + 浅红底），合计行加粗
"""

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .processor import WEEKLY_TARGET, V2_HEADERS, V2_STATUS_ORDER

# —— 复用模板配色 ——
FONT_NAME = "微软雅黑"
HEADER_FILL = PatternFill(patternType="solid", fgColor="D6E0F5")  # accent1 tint0.6 近似
TITLE_FILL = PatternFill(patternType="solid", fgColor="4874CB")   # accent1 原色（标题条）
RED_FILL = PatternFill(patternType="solid", fgColor="FCEEEE")     # 未达标：极浅红底，仅作提示
RED_FONT_COLOR = "C0504D"                                          # 柔和红字
WHITE = "FFFFFF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["客户经理", "今周目标", "预约数", "走访数", "预约完成率", "差值"]
PCT_COLS = {"预约完成率"}


def write_styled_table(df, out_path, title="政企家庭专项走访统计（今周目标 5 户/人）"):
    """把结果 DataFrame 写入 out_path（新建工作簿），套用早会五张表风格。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "家庭专项走访统计"

    ncol = len(HEADERS)
    last_col = get_column_letter(ncol)

    # 第 1 行：标题条
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws["A1"]
    tcell.value = title
    tcell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    tcell.fill = TITLE_FILL
    tcell.alignment = CENTER

    # 第 2 行：表头
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 数据行
    total_row_idx = 1 + 1 + len(df)  # 标题+表头+数据
    for i, (_, row) in enumerate(df.iterrows(), start=3):
        is_total = str(row["客户经理"]) == "合计"
        appt_below = (not is_total) and (int(row["预约数"]) < WEEKLY_TARGET)
        for j, h in enumerate(HEADERS, start=1):
            val = row[h]
            c = ws.cell(row=i, column=j)
            if h in PCT_COLS:
                c.value = float(val)
                c.number_format = "0.0%"
            else:
                c.value = val
            c.font = Font(name=FONT_NAME, size=11, bold=is_total)
            c.alignment = CENTER
            c.border = BORDER
            # 预约数未达标：仅预约完成率列做柔和提示（浅红底 + 柔和红字）
            if appt_below and h == "预约完成率":
                c.fill = RED_FILL
                c.font = Font(name=FONT_NAME, size=11, bold=False, color=RED_FONT_COLOR)

    # 列宽
    widths = {"客户经理": 12, "今周目标": 10, "预约数": 9, "走访数": 9,
              "预约完成率": 13, "差值": 8}
    for j, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 10)
    ws.row_dimensions[1].height = 26

    wb.save(out_path)
    return out_path


V2_TEAM_FILL = PatternFill(patternType="solid", fgColor="E8F0FE")
V2_TOTAL_FILL = PatternFill(patternType="solid", fgColor="FFF4D6")
V2_PIE_COLORS = ["2A78D6", "EB6834", "1BAF7A"]


def _write_v2_summary_sheet(wb, summary_df, title):
    ws = wb.active
    ws.title = "统计汇总"
    last_col = get_column_letter(len(V2_HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER

    for col, header in enumerate(V2_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name=FONT_NAME, size=11, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=3):
        name = str(row["客户经理"])
        is_total = name == "总计"
        is_team = name.endswith("团队合计")
        for col, header in enumerate(V2_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col)
            value = row[header]
            if header == "转化率":
                cell.value = float(value)
                cell.number_format = "0.0%"
            elif header in ("转化积分合计", "转化高套合计"):
                cell.value = float(value)
                cell.number_format = "0.000"
            else:
                cell.value = value
            cell.font = Font(name=FONT_NAME, size=11, bold=is_total or is_team)
            cell.alignment = CENTER
            cell.border = BORDER
            if is_total:
                cell.fill = V2_TOTAL_FILL
            elif is_team:
                cell.fill = V2_TEAM_FILL

    widths = {
        "客户经理": 18, "预约数": 10, "走访数": 10, "转化数": 10,
        "转化积分合计": 16, "转化高套合计": 16, "转化率": 12,
    }
    for col, header in enumerate(V2_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths[header]
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A3"


def _write_v2_pie_sheet(wb, pie_counts):
    ws = wb.create_sheet("商机转化饼图")
    ws["A1"] = "商机转化情况分布"
    ws["A1"].font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = CENTER
    ws.merge_cells("A1:D1")

    group_columns = [("总体", 1), ("党政军团队", 5), ("大企业团队", 9)]
    for group_name, start_col in group_columns:
        ws.cell(row=3, column=start_col, value="商机转化情况")
        ws.cell(row=3, column=start_col + 1, value="数量")
        for col in (start_col, start_col + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        counts = pie_counts.get(group_name, {})
        for row_idx, status in enumerate(V2_STATUS_ORDER, start=4):
            label = ws.cell(row=row_idx, column=start_col, value=status)
            value = ws.cell(row=row_idx, column=start_col + 1, value=counts.get(status, 0))
            for cell in (label, value):
                cell.font = Font(name=FONT_NAME, size=11)
                cell.alignment = CENTER
                cell.border = BORDER

        chart = PieChart()
        labels = Reference(ws, min_col=start_col, min_row=4, max_row=6)
        values = Reference(ws, min_col=start_col + 1, min_row=3, max_row=6)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = group_name
        chart.height = 8
        chart.width = 12
        chart.varyColors = False
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLeaderLines = True
        series = chart.series[0]
        series.data_points = []
        for index, color in enumerate(V2_PIE_COLORS):
            point = DataPoint(idx=index)
            point.graphicalProperties.solidFill = color
            point.graphicalProperties.line.noFill = True
            series.data_points.append(point)
        ws.add_chart(chart, f"{get_column_letter(start_col)}9")

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_v2_workbook(summary_df, pie_counts, out_path, title):
    """生成第二版统计汇总与商机转化饼图工作簿。"""
    wb = Workbook()
    _write_v2_summary_sheet(wb, summary_df, title)
    _write_v2_pie_sheet(wb, pie_counts)
    wb.save(out_path)
    return out_path
