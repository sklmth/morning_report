"""
专项业绩奖励 — 完美一单解析 + 增量统计 + Excel 导出
=======================================================
数据来源：上传的「完美一单」Excel，从「揽装人维度（月累）」sheet 读取：
  · 日期：A10（iloc[9, 0]）
  · 数据行：第6行起（iloc[5:]），列索引 0-based：
      col 4  客户经理名称
      col 5  新增高套
      col 6  存量高套
      col 13 增量积分总分
Excel 导出风格对齐 wecom_notice/excel_export.py（微软雅黑 + 深蓝标题 + 浅蓝表头）。
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from wecom_notice.config import CUSTOMER_MANAGERS

# ── 仅统计非实习期客户经理 ──────────────────────────────────────────────────
_ACTIVE_MANAGERS: set[str] = {
    m["name"] for m in CUSTOMER_MANAGERS if not m.get("exclude_reminder", False)
}

# ── 揽装人维度（月累）sheet 列映射（0-based） ──────────────────────────────
_NAME_COL = 4
_NEW_GAOTAO_COL = 5
_STOCK_GAOTAO_COL = 6
_INC_PTS_COL = 13
_DATE_ROW = 9      # A10 = iloc[9, 0]
_DATA_START_ROW = 5  # 数据从第6行起（iloc[5:]）

# ── Excel 样式常量（复用 wecom_notice/excel_export.py 配色） ──────────────
_FONT = "微软雅黑"
_TITLE_FILL = PatternFill(patternType="solid", fgColor="4874CB")
_HEADER_FILL = PatternFill(patternType="solid", fgColor="D6E0F5")
_ALT_FILL = PatternFill(patternType="solid", fgColor="EEF3FB")
_WHITE_FILL = PatternFill(patternType="solid", fgColor="FFFFFF")
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_WHITE = "FFFFFF"


# ─────────────────────────────────────────────────────────────────────────────
# 一、解析完美一单
# ─────────────────────────────────────────────────────────────────────────────

def _safe_float(val, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        f = float(val)
        return default if (f != f) else f  # NaN check
    except (TypeError, ValueError):
        return default


def _parse_date(raw) -> Optional[str]:
    """将各种格式的日期值统一为 YYYY-MM-DD。"""
    if raw is None:
        return None
    s = str(raw).strip()
    # 纯8位数字  20260806
    m = re.search(r"(\d{8})", s)
    if m:
        d = m.group(1)
        return f"{d[:4]}-{d[4:6]}-{d[6:]}"
    # YYYY-MM-DD 或 YYYY/MM/DD
    m = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        y, mo, dy = m.group(1), m.group(2).zfill(2), m.group(3).zfill(2)
        return f"{y}-{mo}-{dy}"
    # pandas Timestamp
    try:
        dt = pd.Timestamp(raw)
        if not pd.isnull(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def parse_wanmei_excel(file_path: str) -> dict:
    """
    解析完美一单 Excel，提取每位（非实习）客户经理的累计积分和高套数。

    Returns::
        {
            "file_date": "2026-08-06",
            "stats": [
                {"manager_name": "张三", "cumulative_points": 100.0, "cumulative_gaotao": 5.0},
            ]
        }
    Raises:
        ValueError: sheet 不存在或格式不符
    """
    data = pd.read_excel(file_path, sheet_name=None, header=None)
    target_sheet = None
    for sname in data.keys():
        if "揽装人维度" in sname and "月累" in sname:
            target_sheet = sname
            break
    if target_sheet is None:
        raise ValueError(
            f"完美一单缺少「揽装人维度（月累）」sheet，实际：{list(data.keys())}"
        )

    df = data[target_sheet]

    # A10 = 索引 9 行读日期
    file_date: Optional[str] = None
    if len(df) > _DATE_ROW:
        file_date = _parse_date(df.iloc[_DATE_ROW, 0])

    data_rows = df.iloc[_DATA_START_ROW:].reset_index(drop=True)
    stats: list[dict] = []
    seen: set[str] = set()
    for _, row in data_rows.iterrows():
        if len(row) <= _NAME_COL:
            continue
        name = str(row.iloc[_NAME_COL]).strip() if pd.notna(row.iloc[_NAME_COL]) else ""
        if not name or name in ("nan", "客户经理名称", "姓名") or name in seen:
            continue
        if name not in _ACTIVE_MANAGERS:
            continue
        seen.add(name)
        new_gt = _safe_float(row.iloc[_NEW_GAOTAO_COL] if len(row) > _NEW_GAOTAO_COL else 0)
        stk_gt = _safe_float(row.iloc[_STOCK_GAOTAO_COL] if len(row) > _STOCK_GAOTAO_COL else 0)
        pts = _safe_float(row.iloc[_INC_PTS_COL] if len(row) > _INC_PTS_COL else 0)
        stats.append({
            "manager_name": name,
            "cumulative_points": pts,
            "cumulative_gaotao": new_gt + stk_gt,
        })

    return {"file_date": file_date, "stats": stats}


# ─────────────────────────────────────────────────────────────────────────────
# 二、增量统计
# ─────────────────────────────────────────────────────────────────────────────

def compute_incremental_stats(month: str) -> dict:
    """
    计算本月各上传版本的增量数据。

    增量 = 本次累计 − 上次累计（按 manager_name 对齐）。
    第一次上传没有增量（全为 0）。

    Returns::
        {
            "uploads": [...],
            "cumulative": {upload_id: [{"manager_name", "cumulative_points", "cumulative_gaotao"}, ...]},
            "incremental": {upload_id: [{"manager_name", "inc_points", "inc_gaotao"}, ...]},
        }
    """
    from wecom_notice.db import get_performance_uploads, get_performance_stats

    uploads = get_performance_uploads(month)
    if not uploads:
        return {"uploads": [], "cumulative": {}, "incremental": {}}

    cumulative: dict[int, list[dict]] = {}
    incremental: dict[int, list[dict]] = {}
    prev_stats: dict[str, dict] = {}

    for upload in uploads:
        uid = upload["id"]
        rows = get_performance_stats(month, upload_id=uid)

        cumulative[uid] = [
            {"manager_name": r["manager_name"],
             "cumulative_points": r["cumulative_points"],
             "cumulative_gaotao": r["cumulative_gaotao"]}
            for r in rows
        ]

        inc_rows = []
        for r in rows:
            name = r["manager_name"]
            prev = prev_stats.get(name, {"cumulative_points": 0.0, "cumulative_gaotao": 0.0})
            inc_rows.append({
                "manager_name": name,
                "inc_points": r["cumulative_points"] - prev["cumulative_points"],
                "inc_gaotao": r["cumulative_gaotao"] - prev["cumulative_gaotao"],
            })
        incremental[uid] = inc_rows
        prev_stats = {r["manager_name"]: r for r in rows}

    return {"uploads": uploads, "cumulative": cumulative, "incremental": incremental}


# ─────────────────────────────────────────────────────────────────────────────
# 三、Excel 导出（风格对齐企业走访通报）
# ─────────────────────────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int, value=None, bold=False, fill=None,
          align=None, font_size: int = 11, font_color: str = "000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=_FONT, size=font_size, bold=bold, color=font_color)
    c.border = _BORDER
    c.alignment = align or _CENTER
    if fill:
        c.fill = fill
    return c


def export_performance_excel(
    month: str,
    upload_id: int,
    prev_snapshot: "dict | None" = None,
) -> bytes:
    """
    生成专项业绩通报 Excel，返回 bytes（可直接发送给前端下载）。

    包含两个 sheet（积分排名 + 高套排名），各 4 列：
      名次 | 客户经理 | 本月累计 | 新增（vs 上次发奖快照，第一次发奖前显示 "-"）
    仅统计非实习期客户经理（_ACTIVE_MANAGERS）。

    Args:
        month:         月份，格式 "YYYY-MM"
        upload_id:     当前完美一单上传记录 id
        prev_snapshot: get_latest_dispatch_snapshot_map() 的返回值，
                       格式 {manager_name: {cumulative_points, cumulative_gaotao, dispatch_date}}；
                       为 None 表示本月尚未发过奖励，新增列全显示 "-"。
    """
    from wecom_notice.db import get_performance_stats, get_performance_uploads

    # ── 获取当前数据 ─────────────────────────────────────────────────────────
    cur_rows: list[dict] = get_performance_stats(month, upload_id=upload_id)
    uploads = get_performance_uploads(month)
    upload_info = next((u for u in uploads if u["id"] == upload_id), {})
    file_date = upload_info.get("file_date", "")

    # ── 新增列标题 ───────────────────────────────────────────────────────────
    if prev_snapshot:
        sample = next(iter(prev_snapshot.values()), {})
        snap_date = sample.get("dispatch_date", "上次发奖")
        inc_col_label = f"新增{{unit}}（vs {snap_date} 发奖）"
    else:
        inc_col_label = "新增{unit}（暂无发奖记录）"

    # ── 构建排名数据 ─────────────────────────────────────────────────────────
    rows_pts = sorted(cur_rows, key=lambda x: x["cumulative_points"], reverse=True)
    rows_gt  = sorted(cur_rows, key=lambda x: x["cumulative_gaotao"],  reverse=True)

    wb = Workbook()

    def write_sheet(ws, title: str, sorted_rows: list[dict], cum_key: str, unit: str):
        col_headers = [
            "名次", "客户经理",
            f"本月累计{unit}",
            inc_col_label.replace("{unit}", unit),
        ]
        ncol = len(col_headers)
        last_col = get_column_letter(ncol)

        # 标题行
        ws.merge_cells(f"A1:{last_col}1")
        tc = ws["A1"]
        tc.value = title
        tc.font = Font(name=_FONT, size=13, bold=True, color=_WHITE)
        tc.fill = _TITLE_FILL
        tc.alignment = _CENTER

        # 表头行
        for j, h in enumerate(col_headers, 1):
            _cell(ws, 2, j, h, bold=True, fill=_HEADER_FILL, font_size=11)

        # 数据行
        total_inc = 0.0
        for i, row in enumerate(sorted_rows, 1):
            name = row["manager_name"]
            cum_val = row[cum_key]
            fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
            r = i + 2
            _cell(ws, r, 1, i,    fill=fill)
            _cell(ws, r, 2, name, fill=fill, align=_LEFT)
            _cell(ws, r, 3, round(cum_val, 2), fill=fill)
            if prev_snapshot is not None:
                prev = prev_snapshot.get(name, {})
                inc_val = cum_val - prev.get(cum_key, 0.0)
                total_inc += inc_val
                inc_str = f"+{round(inc_val, 2)}" if inc_val > 0 else str(round(inc_val, 2))
            else:
                inc_str = "-"
            _cell(ws, r, 4, inc_str, fill=fill)

        # 合计行
        total_cum = sum(r[cum_key] for r in sorted_rows)
        r_total = len(sorted_rows) + 3
        _cell(ws, r_total, 1, "合计",              bold=True, fill=_HEADER_FILL)
        _cell(ws, r_total, 2, f"{len(sorted_rows)} 人", bold=True, fill=_HEADER_FILL)
        _cell(ws, r_total, 3, round(total_cum, 2), bold=True, fill=_HEADER_FILL)
        if prev_snapshot is not None:
            _cell(ws, r_total, 4,
                  f"+{round(total_inc,2)}" if total_inc >= 0 else str(round(total_inc,2)),
                  bold=True, fill=_HEADER_FILL)
        else:
            _cell(ws, r_total, 4, "-", bold=True, fill=_HEADER_FILL)

        # 列宽
        col_widths = [8, 14, 18, 28]
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.row_dimensions[1].height = 30

    # 积分 sheet
    ws_pts = wb.active
    ws_pts.title = "积分排名"
    write_sheet(ws_pts,
                f"【专项业绩】{month} 积分统计（截至 {file_date}）",
                rows_pts, "cumulative_points", "积分")

    # 高套 sheet
    ws_gt = wb.create_sheet("高套排名")
    write_sheet(ws_gt,
                f"【专项业绩】{month} 高套统计（截至 {file_date}）",
                rows_gt, "cumulative_gaotao", "高套数")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
