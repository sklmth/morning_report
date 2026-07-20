"""建库：documents/ 下的文档 -> 解析 -> 切块 -> 向量化 -> 存入 ChromaDB。

用法：
    python ingest.py            # 处理 documents/ 下所有文档
"""
import sys
from pathlib import Path

import chromadb
from fastembed import TextEmbedding

import config

# ---- 文档解析 ----------------------------------------------------------

def read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def read_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def read_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"[表: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


READERS = {
    ".txt": read_txt, ".md": read_txt,
    ".docx": read_docx,
    ".pdf": read_pdf,
    ".xlsx": read_xlsx,
}


def load_text(path: Path) -> str:
    reader = READERS.get(path.suffix.lower())
    if not reader:
        return ""
    try:
        return reader(path)
    except Exception as e:
        print(f"  ! 解析失败 {path.name}: {e}")
        return ""

# ---- 切块 --------------------------------------------------------------

def chunk_text(text: str, size: int, overlap: int) -> list[str]:
    """按字符切块，块间重叠。先按段落合并，避免切断句子。"""
    text = text.strip()
    if not text:
        return []
    chunks, buf = [], ""
    for para in text.split("\n"):
        para = para.strip()
        if not para:
            continue
        if len(buf) + len(para) <= size:
            buf += para + "\n"
        else:
            if buf:
                chunks.append(buf.strip())
            # 段落本身超长时，硬切
            while len(para) > size:
                chunks.append(para[:size])
                para = para[size - overlap:]
            buf = para + "\n"
    if buf.strip():
        chunks.append(buf.strip())
    return chunks


# ---- 可复用的建库组件（供 CLI 与 API 共用）--------------------------------

SUPPORTED_EXTS = tuple(READERS.keys())


def load_embedder():
    """加载本地 Embedding 模型。"""
    return TextEmbedding(model_name=config.EMBED_MODEL)


def get_client():
    return chromadb.PersistentClient(path=str(config.DB_DIR))


def open_or_create_collection(client):
    """打开集合；不存在则创建（供增量上传时首次建库用）。"""
    try:
        return client.get_collection(config.COLLECTION)
    except Exception:
        return client.create_collection(
            config.COLLECTION, metadata={"hnsw:space": "cosine"})


def _chunk_id(source: str, i: int) -> str:
    # 用完整文件名做前缀，避免不同文档同名 stem 时 id 冲突
    return f"{source}::{i}"


def ingest_file(embedder, coll, path: Path) -> int:
    """把单个文档增量写入向量库。

    若该来源(文件名)已存在，先删除其旧块再写入，保证重复上传时幂等。
    返回写入的文本块数量。
    """
    text = load_text(path)
    chunks = chunk_text(text, config.CHUNK_SIZE, config.CHUNK_OVERLAP)
    source = path.name

    # 先清掉同名来源的旧块（覆盖式更新）
    try:
        coll.delete(where={"source": source})
    except Exception:
        pass

    if not chunks:
        return 0

    ids = [_chunk_id(source, i) for i in range(len(chunks))]
    metas = [{"source": source, "chunk": i} for i in range(len(chunks))]
    vectors = [v.tolist() for v in embedder.embed(chunks)]
    coll.add(documents=chunks, embeddings=vectors, metadatas=metas, ids=ids)
    return len(chunks)


def delete_source(coll, source: str) -> None:
    """从向量库中删除某个来源(文件名)的全部文本块。"""
    coll.delete(where={"source": source})


def list_sources(coll) -> list[dict]:
    """列出向量库中已入库的文档及其块数。"""
    try:
        data = coll.get(include=["metadatas"])
    except Exception:
        return []
    counts: dict[str, int] = {}
    for m in data.get("metadatas") or []:
        src = (m or {}).get("source")
        if src:
            counts[src] = counts.get(src, 0) + 1
    return [{"source": s, "chunks": c} for s, c in sorted(counts.items())]


# ---- 主流程 ------------------------------------------------------------

def main():
    if not config.DOCS_DIR.exists():
        config.DOCS_DIR.mkdir(parents=True)
        print(f"已创建 {config.DOCS_DIR}，请放入文档后重试。")
        return

    files = [p for p in config.DOCS_DIR.rglob("*") if p.suffix.lower() in READERS]
    if not files:
        print(f"{config.DOCS_DIR} 下没有可处理的文档（支持 txt/md/docx/pdf/xlsx）。")
        return

    print(f"加载 Embedding 模型 {config.EMBED_MODEL} ...")
    embedder = load_embedder()

    client = get_client()
    # 全量重建集合，保证幂等
    try:
        client.delete_collection(config.COLLECTION)
    except Exception:
        pass
    coll = client.create_collection(config.COLLECTION, metadata={"hnsw:space": "cosine"})

    total = 0
    for f in files:
        n = ingest_file(embedder, coll, f)
        print(f"  {f.name}: {n} 块")
        total += n

    if not total:
        print("没有提取到任何文本。")
        return

    print(f"完成！共入库 {total} 块，来自 {len(files)} 个文档。")
    print(f"向量库位置: {config.DB_DIR}")


if __name__ == "__main__":
    main()
