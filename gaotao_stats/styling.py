"""把新增/存量高套明细写成带样式的 Excel，风格对齐「早会五张表」。

样式取自 zhengqi_visit_stats/styling.py：
    字体   微软雅黑 11
    标题条 accent1 原色 4874CB + 白字
    表头   浅蓝底 D6E0F5、居中、加粗
    边框   全表细线
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "微软雅黑"
HEADER_FILL = PatternFill(patternType="solid", fgColor="D6E0F5")
TITLE_FILL = PatternFill(patternType="solid", fgColor="4874CB")
WHITE = "FFFFFF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["接入号", "客户经理", "竣工日期", "积分", "高套数"]
COL_WIDTHS = {"接入号": 24, "客户经理": 12, "竣工日期": 13, "积分": 10, "高套数": 10}
NUM_COLS = {"积分", "高套数"}


def _write_sheet(ws, df, title):
    ncol = len(HEADERS)
    last_col = get_column_letter(ncol)

    # 第 1 行：标题条
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws["A1"]
    tcell.value = title
    tcell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    tcell.fill = TITLE_FILL
    tcell.alignment = CENTER

    # 第 2 行：表头
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 数据行
    for i, (_, row) in enumerate(df.iterrows(), start=3):
        for j, h in enumerate(HEADERS, start=1):
            val = row[h]
            c = ws.cell(row=i, column=j)
            if h in NUM_COLS:
                c.value = float(val) if val != "" else 0
            else:
                c.value = "" if val is None else str(val)
            c.font = Font(name=FONT_NAME, size=11)
            c.alignment = CENTER
            c.border = BORDER

    # 合计行
    total_row = 2 + len(df) + 1
    ws.cell(row=total_row, column=1, value="合计")
    ws.cell(row=total_row, column=2, value=f"{len(df)} 户")
    ws.cell(row=total_row, column=4, value=float(df["积分"].sum()) if len(df) else 0)
    ws.cell(row=total_row, column=5, value=float(df["高套数"].sum()) if len(df) else 0)
    for j in range(1, ncol + 1):
        c = ws.cell(row=total_row, column=j)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.alignment = CENTER
        c.border = BORDER
        c.fill = HEADER_FILL

    # 列宽 & 行高
    for j, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTHS.get(h, 12)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A3"


def write_styled_workbook(df_new, df_stock, out_path):
    """写出含「新增高套」「存量高套」两 sheet 的工作簿。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "新增高套"
    _write_sheet(ws1, df_new, "客户经理新增高套清单")

    ws2 = wb.create_sheet("存量高套")
    _write_sheet(ws2, df_stock, "客户经理存量高套清单")

    wb.save(out_path)
    return out_path
