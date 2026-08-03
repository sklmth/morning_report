"""累计统计报表Excel导出，风格对齐早会五张表。

参考 gaotao_stats/styling.py 和 zhengqi_visit_stats/styling.py 的样式。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 复用早会五张表配色
FONT_NAME = "微软雅黑"
HEADER_FILL = PatternFill(patternType="solid", fgColor="D6E0F5")  # 浅蓝表头
TITLE_FILL = PatternFill(patternType="solid", fgColor="4874CB")   # 深蓝标题条
GREEN_FILL = PatternFill(patternType="solid", fgColor="E8F5E9")   # 准时填报（绿色）
YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFF9E6")  # 超时填报（黄色）
RED_FILL = PatternFill(patternType="solid", fgColor="FCEEEE")     # 漏填（红色）
WHITE = "FFFFFF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_sheet(ws, data_list, title, bg_fill, show_dates=True):
    """
    写入单个sheet的统计数据。

    Args:
        ws: worksheet对象
        data_list: 数据列表，每项包含 manager_name, count, dates, details
        title: 标题文本
        bg_fill: 数据行背景色
        show_dates: 是否显示详细日期列表
    """
    if show_dates:
        headers = ["客户经理", "次数", "日期列表", "平均提醒次数"]
        col_widths = [15, 10, 40, 15]
    else:
        headers = ["客户经理", "次数", "最近日期", "平均提醒次数"]
        col_widths = [15, 10, 15, 15]

    ncol = len(headers)
    last_col = get_column_letter(ncol)

    # 第1行：标题条
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws["A1"]
    tcell.value = title
    tcell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    tcell.fill = TITLE_FILL
    tcell.alignment = CENTER

    # 第2行：表头
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 数据行
    total_count = 0
    total_reminders = 0
    row_idx = 3

    for item in data_list:
        manager_name = item["manager_name"]
        count = item["count"]
        dates = item.get("dates", [])
        details = item.get("details", [])

        # 计算平均提醒次数
        if details:
            avg_reminders = sum(d.get("reminder_count", 0) for d in details) / len(details)
        else:
            avg_reminders = 0

        total_count += count
        total_reminders += avg_reminders * count

        # 日期显示
        if show_dates:
            date_str = ", ".join(dates[:10])  # 最多显示10个日期
            if len(dates) > 10:
                date_str += f" ... 还有{len(dates) - 10}个"
        else:
            date_str = dates[0] if dates else ""

        ws.cell(row=row_idx, column=1, value=manager_name)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=date_str)
        ws.cell(row=row_idx, column=4, value=round(avg_reminders, 1))

        for j in range(1, ncol + 1):
            c = ws.cell(row=row_idx, column=j)
            c.font = Font(name=FONT_NAME, size=11)
            c.alignment = LEFT if j == 3 else CENTER
            c.border = BORDER
            c.fill = bg_fill

        row_idx += 1

    # 合计行
    if data_list:
        avg_total_reminders = total_reminders / total_count if total_count > 0 else 0
        ws.cell(row=row_idx, column=1, value="合计")
        ws.cell(row=row_idx, column=2, value=total_count)
        ws.cell(row=row_idx, column=3, value=f"{len(data_list)}位客户经理")
        ws.cell(row=row_idx, column=4, value=round(avg_total_reminders, 1))

        for j in range(1, ncol + 1):
            c = ws.cell(row=row_idx, column=j)
            c.font = Font(name=FONT_NAME, size=11, bold=True)
            c.alignment = LEFT if j == 3 else CENTER
            c.border = BORDER
            c.fill = HEADER_FILL

    # 列宽
    for j, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = width

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A3"


def export_cumulative_stats(stats_data, out_path, date_range=""):
    """
    导出累计统计Excel，包含四个sheet：
    1. 汇总统计
    2. 准时填报（做得好的在前面）
    3. 超时填报
    4. 漏填通报

    Args:
        stats_data: reporter.build_cumulative_statistics()返回的数据
        out_path: 输出Excel路径
        date_range: 日期范围描述（如"2025-01-01至2025-01-31"）
    """
    suffix = f"（{date_range}）" if date_range else ""
    wb = Workbook()

    # Sheet 1: 汇总统计
    ws_summary = wb.active
    ws_summary.title = "汇总统计"
    summary = stats_data["summary"]

    # 标题
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"预约填报累计统计汇总{suffix}"
    title_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    ws_summary.row_dimensions[1].height = 26

    # 汇总数据
    summary_items = [
        ("统计区间", f"{summary['start_date']} 至 {summary['end_date']}"),
        ("统计天数", f"{summary['total_days']} 天"),
        ("总记录数", f"{summary['total_records']} 条"),
        ("", ""),
        ("准时填报", f"{summary['on_time_count']} 次（{summary['on_time_rate']:.1%}）"),
        ("超时填报", f"{summary['overtime_count']} 次（{summary['overtime_rate']:.1%}）"),
        ("漏填次数", f"{summary['missing_count']} 次（{summary['missing_rate']:.1%}）"),
    ]

    for row_idx, (label, value) in enumerate(summary_items, start=2):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

        for col in (1, 2):
            c = ws_summary.cell(row=row_idx, column=col)
            c.font = Font(name=FONT_NAME, size=11, bold=(label != ""))
            c.alignment = LEFT
            c.border = BORDER if label else None
            if row_idx >= 5 and label:
                if "准时" in label:
                    c.fill = GREEN_FILL
                elif "超时" in label:
                    c.fill = YELLOW_FILL
                elif "漏填" in label:
                    c.fill = RED_FILL

    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 35

    # Sheet 2: 准时填报
    ws_on_time = wb.create_sheet("准时填报")
    _write_sheet(
        ws_on_time,
        stats_data["on_time"],
        f"准时填报统计（19:30前完成）{suffix}",
        GREEN_FILL,
        show_dates=True
    )

    # Sheet 3: 超时填报
    ws_overtime = wb.create_sheet("超时填报")
    _write_sheet(
        ws_overtime,
        stats_data["overtime"],
        f"超时填报统计（19:30-23:30）{suffix}",
        YELLOW_FILL,
        show_dates=True
    )

    # Sheet 4: 漏填通报
    ws_missing = wb.create_sheet("漏填通报")
    _write_sheet(
        ws_missing,
        stats_data["missing"],
        f"漏填通报统计（23:30未完成）{suffix}",
        RED_FILL,
        show_dates=True
    )

    wb.save(out_path)
    return out_path
