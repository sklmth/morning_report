"""把统计 DataFrame 写成带样式的 Excel，风格对齐「早会五张表」。

样式要点（取自 assets/早会五张表.xlsx 的模板1）：
    字体      微软雅黑 11
    表头      主题色 accent1(4874CB) + tint 0.6 的浅蓝底、居中
    边框      全表细线
    完成率    百分比显示
    预约数    < 5 户标红（红字 + 浅红底），合计行加粗
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .processor import WEEKLY_TARGET

# —— 复用模板配色 ——
FONT_NAME = "微软雅黑"
HEADER_FILL = PatternFill(patternType="solid", fgColor="D6E0F5")  # accent1 tint0.6 近似
TITLE_FILL = PatternFill(patternType="solid", fgColor="4874CB")   # accent1 原色（标题条）
RED_FILL = PatternFill(patternType="solid", fgColor="FCEEEE")     # 未达标：极浅红底，仅作提示
RED_FONT_COLOR = "C0504D"                                          # 柔和红字
WHITE = "FFFFFF"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADERS = ["客户经理", "今周目标", "预约数", "走访数", "预约完成率", "差值"]
PCT_COLS = {"预约完成率"}


def write_styled_table(df, out_path, title="政企家庭专项走访统计（今周目标 5 户/人）"):
    """把结果 DataFrame 写入 out_path（新建工作簿），套用早会五张表风格。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "家庭专项走访统计"

    ncol = len(HEADERS)
    last_col = get_column_letter(ncol)

    # 第 1 行：标题条
    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws["A1"]
    tcell.value = title
    tcell.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
    tcell.fill = TITLE_FILL
    tcell.alignment = CENTER

    # 第 2 行：表头
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    # 数据行
    total_row_idx = 1 + 1 + len(df)  # 标题+表头+数据
    for i, (_, row) in enumerate(df.iterrows(), start=3):
        is_total = str(row["客户经理"]) == "合计"
        appt_below = (not is_total) and (int(row["预约数"]) < WEEKLY_TARGET)
        for j, h in enumerate(HEADERS, start=1):
            val = row[h]
            c = ws.cell(row=i, column=j)
            if h in PCT_COLS:
                c.value = float(val)
                c.number_format = "0.0%"
            else:
                c.value = val
            c.font = Font(name=FONT_NAME, size=11, bold=is_total)
            c.alignment = CENTER
            c.border = BORDER
            # 预约数未达标：仅预约完成率列做柔和提示（浅红底 + 柔和红字）
            if appt_below and h == "预约完成率":
                c.fill = RED_FILL
                c.font = Font(name=FONT_NAME, size=11, bold=False, color=RED_FONT_COLOR)

    # 列宽
    widths = {"客户经理": 12, "今周目标": 10, "预约数": 9, "走访数": 9,
              "预约完成率": 13, "差值": 8}
    for j, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 10)
    ws.row_dimensions[1].height = 26

    wb.save(out_path)
    return out_path
